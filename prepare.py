import time
import os
import json
import pymysql
import openpyxl
import requests

class Mysql(object):

    def __init__(self):
        self.db = ''
        self.cursor = ''
        self.connect()


    def connect(self):
        self.db = pymysql.connect(host='127.0.0.1', user='root', password='123', db='trademark')
        self.cursor = self.db.cursor()

    def select(self, sql):
        self.cursor.execute(sql)
        data = self.cursor.fetchall()
        # self.cursor.close()
        # self.db.close()
        return data

    def insert(self, sql, args=None):
        self.connect()
        self.cursor.executemany(sql, args=args)
        self.db.commit()
        self.cursor.close()
        self.db.close()

class EarlierStage(object):
    # 1. 读取文件中的TXT，把原始文件存入mysql中。
    # 2. 筛选数据，把筛选出来的数据写入C:\Users\joy\Desktop\商标网项目\批量查询示例文件-天眼查.xlsx

    def __init__(self, issue):
        self.issue = issue
        self.drop = ['商标', '知识产权', '代理', '咨询', '律师', '专利',
                     '品牌', '传媒', '传播', '广告',
                     '法律', '事务所', '会计', '财务', '物流', '互联网', '媒体', '大数据', ')', '（']
        self.mysql = Mysql()
        self.items = None



    def read_txt(self):
        result = []
        path = fr'C:\Users\joy\Desktop\商标网项目\{self.issue}'
        files_txt = os.listdir(path)
        for file in files_txt:
            file_name = os.path.join(path, file)
            fo = open(file_name, 'r', encoding='utf-8')
            data = fo.read().strip()
            fo.close()
            rows = json.loads(data)['rows']
            for row in rows:
                result.append(tuple(row.values()))
        total = len(result)
        start_time = time.time()

        # 存入数据库
        sql = "insert into rawdata values(0,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        self.mysql.insert(sql, args=result)
        print("存入成功。共计{},耗时:{:.2f}".format(total, start_time))

    def write_excel(self, rows):
        # 读取数据库
        sql = "SELECT distinct reg_name from rawdata WHERE CHAR_LENGTH(reg_name) >4 and not reg_name LIKE '%*%';"
        rows = self.mysql.select(sql)
        self.items = rows
        self.write_excel(rows)
        # 写入天眼查示例文件
        path = r'C:\Users\joy\Desktop\商标网项目\批量查询示例文件-天眼查.xlsx'
        book = openpyxl.load_workbook(path)
        sheet = book['Sheet1']
        total, count, c, tr = 0, 0, 0, 3
        for row in rows:
            total += 1
            company = row[0]
            for dp in self.drop:
                if dp in company:
                    break
            else:
                c += 1
                tr += 1
                sheet.cell(tr, 1).value = company
            if tr == 5003 or total == len(rows):
                count += 1
                book.save(rf"C:\Users\joy\Desktop\商标网项目\导天眼查\{str(count)}.xlsx")
                book.close()
                tr = 3
                book = openpyxl.load_workbook(path)
                sheet = book['Sheet1']
        print(f"导入模板成功!共计{c}条数据，{count}个文件")


    def filter(self):
        sql = "SELECT distinct reg_name from rawdata WHERE CHAR_LENGTH(reg_name) >4 and not reg_name LIKE '%*%';"
        self.items = self.mysql.select(sql)
        # print(self.items)
        result = []
        current = 0
        excel_files = r"C:\Users\joy\Desktop\商标网项目\导出的文件"
        files = os.listdir(excel_files)
        for file in files:
            excel_path = os.path.join(excel_files, file)
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb[wb.sheetnames[0]]

            for row in range(4, sheet.max_row + 1):
                phone = sheet.cell(row, 20).value # 电话 排除非电话号码的
                if phone.strip()[0] != '1':
                    continue
                province = sheet.cell(row, 9).value # 省份 排除'北京', '新疆维吾尔自治区', '西藏自治区'
                if province in ['北京市', '新疆维吾尔自治区', '西藏自治区']:
                    continue
                city = sheet.cell(row, 10).value  # 城市 排除绵阳市
                if '绵阳市' in city:
                    continue

                company = sheet.cell(row, 1).value
                company = company.strip(' \n')
                if '曾用名：' in company:
                    company = company.split('曾用名：')[1][:-1]  #公司名称

                if (company,) not in self.items:
                    continue
                current += 1
                person = sheet.cell(row, 2).value
                address = sheet.cell(row, 22).value
                # print("{0:<5} {1:<28} {2:<6} {3:<12} {4:<25}".format(current, company, person, phone, address))
                result.append((self.issue, company, person, phone, address))
            wb.close()

        # 存入excel中。
        SQL = "insert into result (id, issue, reg_name, person, phone, address) values(0,%s,%s,%s,%s,%s)"
        self.mysql.insert(SQL, args=result)
        print(f"存入mysql成功！共计{len(result)}条")

class CompanyInfo(object):

    def __init__(self):
        self.headers = {
            'Host': 'api9.tianyancha.com',
            'content-type': 'application/json',
            'Authorization': Authorization,
            'X-AUTH-TOKEN': AUTHTOKEN,
            'version': 'TYC-Web',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36 Edg/111.0.1661.54',
        }

    def search_company_id(self, params):
        url = f'https://capi.tianyancha.com/cloud-tempest/search/suggest/v3?'
        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "Host": "capi.tianyancha.com",
            "Origin": "https://www.tianyancha.com",
            "Pragma": "no-cache",
            "Referer": "https://www.tianyancha.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36 Edg/95.0.1020.53",
        }
        data = {"keyword":params}
        while True:
            try:
                page_text = requests.post(url=url, headers=headers, json=data, timeout=5).json()
                if page_text['data']:
                    args = page_text['data'][0]['id']
                    return args
                return 'exit'
            except:
                print(f'error: {params}')
                return 'exit'
                time.sleep(1)

    # 公司基本信息
    def base_info(self, reg_num):
        url = f'https://api9.tianyancha.com/services/v3/t/common/baseinfoV5ForApp/{reg_num}'
        response = requests.get(url, headers=self.headers, timeout=8).json()
        return response

    def parse_base_info(self, response):
        r = response['data']
        try:
            isamephone = r['phoneSourceList'][0]['companyTotalStr'] # 是否同电话企业
        except:
            isamephone = 1
        try:
            company_count = r['legalInfo']['companyNum'] # 任职企业数量
        except:
            company_count = 0
        return (isamephone, company_count)

    # 专利信息：商标 商标文书 作品著作 专利数量
    def pattern_info(self, reg_num):
        url = f'https://api9.tianyancha.com/services/v3/expanse/allCountV3?id={reg_num}'
        response = requests.get(url, headers=self.headers, timeout=8).json()
        return response

    def parse_pattern_info(self, response):
        r = response['data']
        tm_count = r['tmCount'] # 商标信息个数
        doc_count = str(r['trademarkDocCount']) # 商标文书个数
        pattern_count = r['patentCountV4'] # 专利信息个数
        works_count = str(r['copyrightWorks'])  # 作品著作权
        return tm_count, doc_count, pattern_count, works_count

    # 同电话企业的法人是否一致
    def person_is_same(self, phone):
        url = 'https://capi.tianyancha.com/cloud-other-information/search/web/samePhoneCompany'
        headers = {
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Content-Length': '48',
            'Content-Type': 'application/json; charset=UTF-8',
            'Host': 'capi.tianyancha.com',
            'Origin': 'https://www.tianyancha.com',
            'Pragma': 'no-cache',
            'Referer': 'https://www.tianyancha.com/',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Google Chrome";v="98"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36',
            'version': 'TYC-Web',
            'X-AUTH-TOKEN': AUTHTOKEN,
            'X-TYCID': TYCID,
        }
        data = {"word": phone, "pageSize": 10, "pageNum": 1}
        while True:
            try:
                response = requests.post(url, headers=headers, json=data, timeout=6)
                if response.status_code == 200:
                    break
            except Exception as e:
                print(e)
        rep = response.json()
        tmp = None
        for item in rep['data']['items']:
            name = item.get('firstPositionValue')
            if name != tmp and tmp is not None:
                return False
            tmp = name
        return True


class ExcelImage(object):
    pass


    def get_company(self, excel_file):
        ls_company = []
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb[wb.sheetnames[0]]
        rows = sheet.max_row
        for row in list(range(rows, 1, -1)):
            value = sheet.cell(row, 1).value
            if value is None:
                break
            ls_company.append(value)
        wb.close()
        return tuple(ls_company)

    def readImg(self):
        ml = Mysql()
        file_path = r'C:\Users\joy\Desktop\商标网项目\1832\image'
        ls = os.listdir(file_path)
        for i in ls:
            excel_file = os.path.join(file_path, i)
            tup_company = self.get_company(excel_file)
            sql = "select reg_name, reg_num from rawdata where reg_name in {}".format(tup_company)
            data = ml.select(sql)
            print(data)



if __name__ == '__main__':

    # *****************************
    Authorization = '0###oo34J0Vvfd7C0D56q7YDdtM8tCao###1679906796647###8348775f139132664ff59fd94b205965'
    AUTHTOKEN = 'eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxODYwODE2MjMwMCIsImlhdCI6MTY3OTg5MjA2NywiZXhwIjoxNjgyNDg0MDY3fQ.OCbVlvHj0_qxWOS0q_OWwR9KxxCZHscCOwCIit0eZV--BKTLQJmp2AOhA8BNN0a3Mmy0am2FvSwu72IZDiiOSA'
    TYCID = '05a6a590b32611edbd3563314b2a71a4'
    issue = '1832'
    # *****************************

    # stage = EarlierStage(issue)
    # r = stage.filter()

    # 实例化对象
    mysql = Mysql()
    # 连接数据库
    mysql.connect()
    # 查询数据
    SQL = "select id, reg_name, phone from result where issue={} and isnull(link)".format(issue)
    data = mysql.select(SQL)
    length = len(data)
    ctm = CompanyInfo()
    count = 0


    for _id, reg_name, phone in data:
        count += 1
        # 公司ID
        try:
            reg_id = ctm.search_company_id(reg_name)
        except:
            continue
        if reg_id == 'exit':
            continue
        print(reg_name, reg_id)
        # 基础信息
        try:
            rep_base = ctm.base_info(reg_id)
            # print(rep_base)
            params_one = ctm.parse_base_info(rep_base)
            isphone, company_count = params_one[0], params_one[1]
            # print(isphone, company_count)
        except Exception as e:
            print(e)
            continue

        try:
            cy = int(company_count)
        except:
            cy = company_count[:-1]

        try:
            sp = int(isphone)
        except:
            sp = isphone[:-1]

        if int(cy) > 3 or int(sp) > 3:  # 在任企业数超过2个或者同电话企业超过3个
            params_two = (None, None, None, None)

        else:  # 如果任职企业数不超过2个
            if 1 < int(sp) < 4:
                person_is_ = ctm.person_is_same(phone)
                if person_is_:
                    try:
                        # 商标信息
                        rep_tm = ctm.pattern_info(reg_id)
                        params_two = ctm.parse_pattern_info(rep_tm)
                    except:
                        continue
                else:
                    params_two = (None, None, None, None)
            else:
                try:
                    # 商标信息
                    rep_tm = ctm.pattern_info(reg_id)
                    params_two = ctm.parse_pattern_info(rep_tm)
                except:
                    continue
        args_ = params_one + params_two + (reg_id,)
        print(f'{count}/{length} {args_}')

        # 插入数据
        sql = "UPDATE result set isphone=%s, reg_count=%s, tm_count=%s, doc_count=%s, pattern_count=%s, works_count=%s, link=%s where id={}".format(_id)
        try:
            mysql.cursor.execute(sql, args=args_)
            mysql.db.commit()
            print('{:*^50}'.format('插入成功'))
        except:
            print('mysql 出错')
            pass
    mysql.db.close()
    mysql.cursor.close()



